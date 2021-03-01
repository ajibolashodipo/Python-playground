import openpyxl

# load existing workbook
book = openpyxl.load_workbook('dogs.xlsx')
sheet = book.active

totalNumDogs = 0
# iterate through rows starting from the second
for row in sheet.iter_rows(min_row=2):
    # print values contained in cells in all the rows
    for cell in row:
        print(cell.value)
    #     capture specific value
    totalNumDogs += row[1].value
    # modify cell value
    row[1].value += 10
    print("New value= ", row[1].value)

print("Total dogs seen today: ", totalNumDogs)

# add new record to spreadsheet
newRow = ("Dingo", 12)
newRow2 = ("Gerbbbbb", 12)

# append record to sheet
sheet.append(newRow)
sheet.append(newRow2)

# save modified sheet
book.save("dogsv1.xlsx")

# create a new spreadsheet
newBook = openpyxl.Workbook()
newSheet = newBook.active
firstRow = ("Name", "Age", "Fav Colour")
newSheet.append(firstRow)
newBook.save("newBook.xlsx")
